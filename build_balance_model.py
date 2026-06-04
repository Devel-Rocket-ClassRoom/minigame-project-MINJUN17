import re, glob, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATAS = r"Assets/_project/Datas"

def field(text, name):
    m = re.search(rf"^\s*{re.escape(name)}:\s*(.+)$", text, re.M)
    if not m: return None
    v = m.group(1).strip()
    return v

def num(v):
    if v is None or v == "": return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except ValueError:
        return v

def load(path, fields):
    t = open(path, encoding="utf-8", errors="ignore").read()
    return {f: num(field(t, f)) for f in fields}

# ---------- styles ----------
ARIAL = "Arial"
HDR_FILL = PatternFill("solid", start_color="305496")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=ARIAL, bold=True, size=14, color="305496")
INPUT_FONT = Font(name=ARIAL, color="0000FF")          # 파랑 = 입력값
LINK_FONT  = Font(name=ARIAL, color="008000")          # 초록 = 시트간 링크
CALC_FONT  = Font(name=ARIAL, color="000000")          # 검정 = 계산
INPUT_FILL = PatternFill("solid", start_color="FFFF00")
BASE_FONT  = Font(name=ARIAL)
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WON = '#,##0'
PCT = '0.0%'

def style_header(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

def autow(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

wb = Workbook()

# ================= MENU =================
ws = wb.active; ws.title = "메뉴"
ws["A1"] = "메뉴 (수익원)"; ws["A1"].font = TITLE_FONT
hdr = ["메뉴", "가격(₩)", "원가(₩)", "마진(₩)", "마진율", "수요weight", "해금비(만족도)"]
ws.append([]); ws.append(hdr); style_header(ws, 3, len(hdr))
menu_files = sorted(glob.glob(f"{DATAS}/MenuData/*.asset"))
r = 4
for f in menu_files:
    d = load(f, ["price","cost","spawnWeight","satisfactionUnlock"])
    name = os.path.splitext(os.path.basename(f))[0]
    ws.cell(r,1,name).font = BASE_FONT
    ws.cell(r,2,d["price"]).number_format = WON
    ws.cell(r,3,d["cost"]).number_format = WON
    ws.cell(r,4,f"=B{r}-C{r}").number_format = WON          # 마진 = 가격-원가
    ws.cell(r,5,f"=IF(B{r}=0,0,D{r}/B{r})").number_format = PCT
    ws.cell(r,6,d["spawnWeight"])
    ws.cell(r,7,d["satisfactionUnlock"]).number_format = WON
    for c in range(1,8): ws.cell(r,c).border = BORDER
    r += 1
menu_last = r-1
# 평균 행
ws.cell(r,1,"평균").font = Font(name=ARIAL, bold=True)
ws.cell(r,4,f"=AVERAGE(D4:D{menu_last})").number_format = WON
ws.cell(r,5,f"=AVERAGE(E4:E{menu_last})").number_format = PCT
for c in (1,4,5): ws.cell(r,c).font = Font(name=ARIAL, bold=True)
menu_avg_marginpct = f"메뉴!E{r}"
autow(ws, [18,12,12,12,10,12,14])

# ================= CUSTOMER =================
ws = wb.create_sheet("손님")
ws["A1"] = "손님 (결제력)"; ws["A1"].font = TITLE_FONT
hdr = ["손님","지갑(₩)","시작만족","인내심(초)","식사+/초","대기-/초","weight","시작해금"]
ws.append([]); ws.append(hdr); style_header(ws,3,len(hdr))
cust_files = sorted(glob.glob(f"{DATAS}/Customer/*.asset"))
r = 4; wallet_rows = []
for f in cust_files:
    d = load(f, ["wallet","baseSatisfaction","patience","eatGainRate","waitPenaltyRate","spawnWeight","unlockedFromStart"])
    name = os.path.splitext(os.path.basename(f))[0]
    isDT = "DT" in name
    ws.cell(r,1,name).font = BASE_FONT
    if d["wallet"] is not None:
        ws.cell(r,2,d["wallet"]).number_format = WON
        if not isDT: wallet_rows.append(r)
    else:
        ws.cell(r,2,"(DT/배달)")
    ws.cell(r,3,d["baseSatisfaction"])
    ws.cell(r,4,d["patience"])
    ws.cell(r,5,d["eatGainRate"])
    ws.cell(r,6,d["waitPenaltyRate"])
    ws.cell(r,7,d["spawnWeight"])
    ws.cell(r,8,"O" if d["unlockedFromStart"]==1 else "X")
    for c in range(1,9): ws.cell(r,c).border = BORDER
    r += 1
# 평균 지갑(매장손님만)
avg_wallet_cell = None
ws.cell(r,1,"평균 지갑(매장)").font = Font(name=ARIAL, bold=True)
if wallet_rows:
    rng = ",".join(f"B{x}" for x in wallet_rows)
    ws.cell(r,2,f"=AVERAGE({rng})").number_format = WON
    ws.cell(r,2).font = Font(name=ARIAL, bold=True)
    avg_wallet_cell = f"손님!B{r}"
autow(ws, [16,12,10,12,10,10,9,10])

# ================= STAFF =================
ws = wb.create_sheet("직원")
ws["A1"] = "직원"; ws["A1"].font = TITLE_FONT
hdr = ["직원","고용비(₩)","월급(₩)","이동속도","속도배수","친절","배달시간"]
ws.append([]); ws.append(hdr); style_header(ws,3,len(hdr))
order = ["Cook_Junior","Cook_Senior","Cook_Manager","Server_Junior","Server_Senior","Server_Manager","Rider_Junior","Rider_Senior","Rider_Manager"]
r = 4
for name in order:
    f = f"{DATAS}/StaffData/{name}.asset"
    if not os.path.exists(f): continue
    d = load(f, ["hireCost","salary","moveSpeed","speedMultiplier","kindness","deliveryTime"])
    ws.cell(r,1,name).font = BASE_FONT
    ws.cell(r,2,d["hireCost"]).number_format = WON
    ws.cell(r,3,d["salary"]).number_format = WON
    ws.cell(r,4,d["moveSpeed"]); ws.cell(r,5,d["speedMultiplier"])
    ws.cell(r,6,d["kindness"]); ws.cell(r,7,d["deliveryTime"])
    for c in range(1,8): ws.cell(r,c).border = BORDER
    r += 1
autow(ws, [16,12,12,10,10,8,10])

# ================= FURNITURE =================
ws = wb.create_sheet("가구")
ws["A1"] = "가구 / 조리도구"; ws["A1"].font = TITLE_FONT
hdr = ["항목","설치비(₩)","해금비(만족도)","사용시간(초)"]
ws.append([]); ws.append(hdr); style_header(ws,3,len(hdr))
furn_files = (sorted(glob.glob(f"{DATAS}/FurnitureData/*.asset")) +
              sorted(glob.glob(f"{DATAS}/FurnitureData/CookingToolData/*.asset")) +
              sorted(glob.glob(f"{DATAS}/FurnitureData/Toilet/*.asset")))
r = 4
for f in furn_files:
    name = os.path.splitext(os.path.basename(f))[0]
    if name == "LayoutData": continue
    d = load(f, ["purchaseCost","satisfactionUnlock","usingDuration"])
    ws.cell(r,1,name).font = BASE_FONT
    ws.cell(r,2,d["purchaseCost"]).number_format = WON
    ws.cell(r,3,d["satisfactionUnlock"]).number_format = WON
    ws.cell(r,4,d["usingDuration"] if d["usingDuration"] is not None else "-")
    for c in range(1,5): ws.cell(r,c).border = BORDER
    r += 1
autow(ws, [18,12,14,12])

# ================= GATES =================
ws = wb.create_sheet("성장게이트")
ws["A1"] = "맵확장 / 마케팅"; ws["A1"].font = TITLE_FONT
ws.append([]); ws.append(["확장","비용(₩)"]); style_header(ws,3,2)
exp = [("Stage1_DTUnlock","DT"),("Stage2_Floor2Hall","2층 홀"),("Stage3_Floor2Toilet","화장실")]
r=4; gate_cost_cell={}
for fn,label in exp:
    d = load(f"{DATAS}/ExpansionData/{fn}.asset", ["unlockCost"])
    ws.cell(r,1,label).font=BASE_FONT
    ws.cell(r,2,d["unlockCost"]).number_format=WON
    for c in (1,2): ws.cell(r,c).border=BORDER
    gate_cost_cell[label]=f"성장게이트!B{r}"
    r+=1
r+=1
ws.cell(r,1,"마케팅").font=HDR_FONT; ws.cell(r,1).fill=HDR_FILL
ws.cell(r,2,"boost").font=HDR_FONT; ws.cell(r,2).fill=HDR_FILL
ws.cell(r,3,"만족도비").font=HDR_FONT; ws.cell(r,3).fill=HDR_FILL
ws.cell(r,4,"기간(달)").font=HDR_FONT; ws.cell(r,4).fill=HDR_FILL
for c in range(1,5): ws.cell(r,c).alignment=Alignment(horizontal="center")
r+=1
for fn,label in [("FlyerAd","전단지"),("SNSAd","SNS"),("TVAd","TV")]:
    d=load(f"{DATAS}/MarketingData/{fn}.asset",["spawnBoost","satisfactionCost","durationMonths"])
    ws.cell(r,1,label).font=BASE_FONT
    ws.cell(r,2,d["spawnBoost"]); ws.cell(r,3,d["satisfactionCost"]).number_format=WON
    ws.cell(r,4,d["durationMonths"])
    for c in range(1,5): ws.cell(r,c).border=BORDER
    r+=1
autow(ws,[16,12,12,10])

# ================= MODEL =================
ws = wb.create_sheet("모델", 0)
ws["A1"] = "💰 하루 순이익 모델"; ws["A1"].font = TITLE_FONT
ws["A2"] = "노랑칸(파란글씨)만 바꾸면 전부 자동 계산됩니다"; ws["A2"].font = Font(name=ARIAL, italic=True, size=9, color="808080")

def label(r, txt, bold=False):
    c = ws.cell(r,1,txt); c.font = Font(name=ARIAL, bold=bold); c.alignment=Alignment(horizontal="left")
def inp(r, val, fmt=None, note=None):
    c = ws.cell(r,2,val); c.font = INPUT_FONT; c.fill = INPUT_FILL; c.border=BORDER
    if fmt: c.number_format=fmt
    if note: ws.cell(r,3,note).font=Font(name=ARIAL,size=9,color="808080")
def calc(r, formula, fmt=None, link=False, note=None, bold=False):
    c = ws.cell(r,2,formula); c.font = (LINK_FONT if link else CALC_FONT)
    if bold: c.font = Font(name=ARIAL, bold=True, color=("008000" if link else "000000"))
    c.border=BORDER
    if fmt: c.number_format=fmt
    if note: ws.cell(r,3,note).font=Font(name=ARIAL,size=9,color="808080")

ws.cell(4,1,"입력 (가정)").font=HDR_FONT; ws.cell(4,1).fill=HDR_FILL
ws.cell(4,2," ").fill=HDR_FILL; ws.cell(4,3," ").fill=HDR_FILL
label(5,"영업시간 (초/일)"); inp(5,120,WON,"(24-8)시 × 7.5초/시간 = 2분")
label(6,"평균 스폰간격 (초)"); inp(6,20,None,"(min10+max30)/2")
label(7,"마케팅 배수"); inp(7,1.0,'0.00',"마케팅 0개=1.0, 3개=약1.64")
label(8,"좌석 수"); inp(8,3,None,"동시 수용 손님 수")
label(9,"손님 1명 평균 점유시간 (초)"); inp(9,30,None,"주문+조리+식사+퇴장")
label(10,"하루 인건비 합계 (₩)"); inp(10,1900,WON,"현재 고용 직원 월급 합 (1달=1일)")

ws.cell(12,1,"계산").font=HDR_FONT; ws.cell(12,1).fill=HDR_FILL
ws.cell(12,2," ").fill=HDR_FILL; ws.cell(12,3," ").fill=HDR_FILL
label(13,"유효 스폰간격 (초)"); calc(13,"=B6/B7",'0.0',note="스폰간격 ÷ 마케팅배수")
label(14,"이론상 손님/일"); calc(14,"=B5/B13",'0.0',note="좌석 무제한 가정")
label(15,"처리한계 손님/일"); calc(15,"=B8*B5/B9",'0.0',note="좌석×영업시간 ÷ 점유시간")
label(16,"실제 손님/일", True); calc(16,"=MIN(B14,B15)",'0.0',bold=True,note="둘 중 작은 값(병목)")
label(17,"평균 지갑 (매장)"); calc(17,f"={avg_wallet_cell}",WON,link=True,note="← 손님 시트")
label(18,"평균 마진율"); calc(18,f"={menu_avg_marginpct}",PCT,link=True,note="← 메뉴 시트")
label(19,"손님당 순이익 (₩)"); calc(19,"=B17*B18",WON,note="지갑 전액 결제 가정")
label(20,"하루 총이익 (₩)", True); calc(20,"=B16*B19",WON,bold=True)
label(21,"하루 인건비 (₩)"); calc(21,"=B10",WON,note="입력값")
label(22,"하루 순이익 (₩)", True)
c=ws.cell(22,2,"=B20-B21"); c.font=Font(name=ARIAL,bold=True,size=12,color="C00000"); c.number_format=WON; c.border=BORDER

ws.cell(24,1,"성장 게이트까지 걸리는 일수").font=HDR_FONT; ws.cell(24,1).fill=HDR_FILL
ws.cell(24,2," ").fill=HDR_FILL; ws.cell(24,3," ").fill=HDR_FILL
label(25,"DT 해금"); calc(25,f"=IF($B$22<=0,\"적자\",{gate_cost_cell['DT']}/$B$22)",'0.0',note="비용 ÷ 하루순이익")
label(26,"2층 해금"); calc(26,f"=IF($B$22<=0,\"적자\",{gate_cost_cell['2층 홀']}/$B$22)",'0.0')
label(27,"화장실 해금"); calc(27,f"=IF($B$22<=0,\"적자\",{gate_cost_cell['화장실']}/$B$22)",'0.0')

autow(ws,[26,16,34])

# move 모델 first already; reorder sheets
wb.move_sheet("모델", -wb.sheetnames.index("모델"))
wb.save("Balance_Model.xlsx")
print("saved Balance_Model.xlsx")
