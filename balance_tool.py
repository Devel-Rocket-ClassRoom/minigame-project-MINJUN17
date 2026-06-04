#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
balance_tool.py — 메뉴/조리도구/손님/직원/가구/성장게이트 밸런스 양방향 관리 도구

  python balance_tool.py export            # 에셋 → Balance_Model.xlsx  (현재 값 끌어오기)
  python balance_tool.py import            # Balance_Model.xlsx → 에셋  (수정값 반영)
  python balance_tool.py import --dry-run  # 미리보기 (파일은 건드리지 않음)

워크플로
  1) export 로 최신 에셋 값을 엑셀로 내려받는다
  2) 엑셀에서 '입력' 칸만 고친다 (회색 마진/마진율 등 계산열은 자동, 손대지 말 것)
       - '조리도구' 칸은 셀 드롭다운에서 선택 (잘못된 이름 방지)
  3) import 로 에셋에 반영한다  →  반드시 git diff 로 확인 후 커밋

핵심 원칙: '에셋이 정답'. export 는 엑셀을 통째로 새로 쓴다.
           엑셀에서 편집 중이면 먼저 import 한 뒤 export 할 것 (안 그러면 편집분이 날아감).
"""

import re, glob, os, sys, argparse

# Windows 콘솔(cp949)에서도 한글/이모지 출력 깨지지 않게
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE  = os.path.dirname(os.path.abspath(__file__))
DATAS = os.path.join(HERE, "Assets", "_project", "Datas")
XLSX  = os.path.join(HERE, "Balance_Model.xlsx")

# ──────────────────────────────────────────────────────────────────────────
#  공통 헬퍼
# ──────────────────────────────────────────────────────────────────────────
def read(path):
    return open(path, encoding="utf-8", errors="ignore").read()

def field(text, name):
    """에셋 텍스트에서 '  name: 값' 한 줄의 값을 문자열로."""
    m = re.search(rf"^\s*{re.escape(name)}:\s*(.+)$", text, re.M)
    return m.group(1).strip() if m else None

def num(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except ValueError:
        return v

def set_line(text, name, value):
    """'  name: ...' 한 줄을 value 로 교체. (들여쓰기 2칸 = 최상위 필드만)
       반환: (새 텍스트, 교체된 개수)."""
    return re.subn(rf"^(  {re.escape(name)}: ).*$", lambda m: m.group(1) + str(value),
                   text, count=1, flags=re.M)

def fmt_num(kind, raw):
    """엑셀 값 → 에셋에 쓸 문자열."""
    if kind == "int":
        return str(int(round(float(raw))))
    if kind == "float":
        return "%g" % float(raw)      # 1.0→'1', 0.5→'0.5', 1.5→'1.5'
    if kind == "bool01":
        s = str(raw).strip().lower()
        return "1" if s in ("o", "1", "true", "y", "yes") else "0"
    return str(raw)

def parse_guid(ref):
    if not ref:
        return None
    m = re.search(r"guid:\s*([0-9a-f]+)", ref)
    return m.group(1) if m else None

def tool_maps():
    """조리도구 name<->guid 매핑 (.meta 에서 동적으로 읽음)."""
    name2guid, guid2name = {}, {}
    for meta in sorted(glob.glob(f"{DATAS}/FurnitureData/CookingToolData/*.asset.meta")):
        name = os.path.basename(meta)[:-len(".asset.meta")]
        m = re.search(r"^guid:\s*([0-9a-f]+)", read(meta), re.M)
        if m:
            name2guid[name] = m.group(1)
            guid2name[m.group(1)] = name
    return name2guid, guid2name

def asset_index():
    """파일이름(확장자 제외) → 풀경로. (Datas 하위 모든 .asset)"""
    idx = {}
    for f in glob.glob(f"{DATAS}/**/*.asset", recursive=True):
        idx[os.path.splitext(os.path.basename(f))[0]] = f
    return idx

# ──────────────────────────────────────────────────────────────────────────
#  스키마: 시트 ↔ 에셋 필드  (export/import 가 동일하게 참조)
#    헤더문자열 → (에셋필드, 종류)   ※ 헤더문자열은 export 와 정확히 일치해야 함
# ──────────────────────────────────────────────────────────────────────────
IMPORT_SCHEMA = {
    "메뉴":   ("메뉴", {
        "조리도구":       ("tool", "tool"),
        "가격(₩)":        ("price", "int"),
        "원가(₩)":        ("cost", "int"),
        "수요weight":     ("spawnWeight", "float"),
        "해금비(만족도)": ("satisfactionUnlock", "int"),
    }),
    "손님":   ("손님", {
        "지갑(₩)":   ("wallet", "int"),
        "시작만족":  ("baseSatisfaction", "int"),
        "인내심(초)": ("patience", "int"),
        "식사+/초":  ("eatGainRate", "int"),
        "대기-/초":  ("waitPenaltyRate", "int"),
        "식사속도":  ("eatSpeed", "int"),
        "이동속도":  ("moveSpeed", "float"),
        "weight":    ("spawnWeight", "float"),
        "시작해금":  ("unlockedFromStart", "bool01"),
    }),
    "직원":   ("직원", {
        "고용비(₩)": ("hireCost", "int"),
        "월급(₩)":   ("salary", "int"),
        "이동속도":  ("moveSpeed", "float"),
        "속도배수":  ("speedMultiplier", "float"),
        "친절":      ("kindness", "int"),
        "배달시간":  ("deliveryTime", "int"),
    }),
    "가구":   ("항목", {
        "설치비(₩)":      ("purchaseCost", "int"),
        "해금비(만족도)": ("satisfactionUnlock", "int"),
        "사용시간(초)":   ("usingDuration", "float"),
    }),
    "맵확장": ("확장", {
        "비용(₩)": ("unlockCost", "int"),
    }),
    "마케팅": ("마케팅", {
        "boost":     ("spawnBoost", "float"),
        "만족도비":  ("satisfactionCost", "int"),
        "기간(달)":  ("durationMonths", "int"),
    }),
}

# ══════════════════════════════════════════════════════════════════════════
#  EXPORT  (에셋 → 엑셀)
# ══════════════════════════════════════════════════════════════════════════
def export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import ColorScaleRule

    lock = os.path.join(os.path.dirname(XLSX), "~$" + os.path.basename(XLSX))
    if os.path.exists(lock):
        print(f"⚠ {os.path.basename(XLSX)} 가 엑셀에서 열려 있는 것 같습니다. 닫고 다시 실행하세요.")
        return 1

    name2guid, guid2name = tool_maps()

    ARIAL = "Arial"
    HDR_FILL   = PatternFill("solid", start_color="305496")
    HDR_FONT   = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(name=ARIAL, bold=True, size=14, color="305496")
    INPUT_FONT = Font(name=ARIAL, color="0000FF")     # 파랑 = 입력값
    LINK_FONT  = Font(name=ARIAL, color="008000")     # 초록 = 시트간 링크
    CALC_FONT  = Font(name=ARIAL, color="000000")     # 검정 = 계산
    CALC_FILL  = PatternFill("solid", start_color="F2F2F2")  # 회색 = 자동계산(손대지마)
    INPUT_FILL = PatternFill("solid", start_color="FFFF00")
    BASE_FONT  = Font(name=ARIAL)
    thin = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    WON, PCT = '#,##0', '0.0%'

    def hstyle(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill, cell.font = HDR_FILL, HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

    def autow(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    def put_header(ws, hdr, row=3):   # 헤더를 고정 행에 명시적으로 (append 의 행 밀림 방지)
        for i, h in enumerate(hdr, 1):
            ws.cell(row, i, h)
        hstyle(ws, row, len(hdr))

    def load(path, fields):
        t = read(path)
        return {f: num(field(t, f)) for f in fields}

    wb = Workbook()

    # ───────────── 메뉴 (조리도구 매칭 + 가격/원가/마진) ─────────────
    ws = wb.active; ws.title = "메뉴"
    ws["A1"] = "메뉴 (수익원) — 조리도구 매칭"; ws["A1"].font = TITLE_FONT
    ws["A2"] = "파랑=입력 / 회색=자동계산(손대지 마세요) / '조리도구'는 드롭다운 선택"
    ws["A2"].font = Font(name=ARIAL, italic=True, size=9, color="808080")
    hdr = ["메뉴", "조리도구", "가격(₩)", "원가(₩)", "마진(₩)", "마진율", "수요weight", "해금비(만족도)"]
    put_header(ws, hdr)
    r = 4
    for f in sorted(glob.glob(f"{DATAS}/MenuData/*.asset")):
        t = read(f)
        name = os.path.splitext(os.path.basename(f))[0]
        tool = guid2name.get(parse_guid(field(t, "tool")), "")
        d = {k: num(field(t, k)) for k in ("price", "cost", "spawnWeight", "satisfactionUnlock")}
        ws.cell(r, 1, name).font = BASE_FONT
        ws.cell(r, 2, tool).font = INPUT_FONT; ws.cell(r, 2).fill = INPUT_FILL
        ws.cell(r, 3, d["price"]).number_format = WON;  ws.cell(r, 3).font = INPUT_FONT; ws.cell(r, 3).fill = INPUT_FILL
        ws.cell(r, 4, d["cost"]).number_format = WON;   ws.cell(r, 4).font = INPUT_FONT; ws.cell(r, 4).fill = INPUT_FILL
        ws.cell(r, 5, f"=C{r}-D{r}").number_format = WON;            ws.cell(r, 5).fill = CALC_FILL
        ws.cell(r, 6, f"=IF(C{r}=0,0,E{r}/C{r})").number_format = PCT; ws.cell(r, 6).fill = CALC_FILL
        ws.cell(r, 7, d["spawnWeight"]).font = INPUT_FONT; ws.cell(r, 7).fill = INPUT_FILL
        ws.cell(r, 8, d["satisfactionUnlock"]).number_format = WON; ws.cell(r, 8).font = INPUT_FONT; ws.cell(r, 8).fill = INPUT_FILL
        for c in range(1, 9): ws.cell(r, c).border = BORDER
        r += 1
    last = r - 1
    ws.cell(r, 1, "평균").font = Font(name=ARIAL, bold=True)
    ws.cell(r, 5, f"=AVERAGE(E4:E{last})").number_format = WON
    ws.cell(r, 6, f"=AVERAGE(F4:F{last})").number_format = PCT
    for c in (1, 5, 6): ws.cell(r, c).font = Font(name=ARIAL, bold=True)
    menu_avg_marginpct = f"메뉴!F{r}"
    # 조리도구 드롭다운
    if name2guid:
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(sorted(name2guid)), allow_blank=True)
        ws.add_data_validation(dv); dv.add(f"B4:B{last}")
    # 마진율 색 스케일 (이상치 한눈에)
    ws.conditional_formatting.add(f"F4:F{last}",
        ColorScaleRule(start_type="num", start_value=0.4, start_color="F8696B",
                       mid_type="num", mid_value=0.7, mid_color="FFEB84",
                       end_type="num", end_value=0.9, end_color="63BE7B"))
    autow(ws, [16, 16, 11, 11, 11, 9, 11, 14])

    # ───────────── 손님 ─────────────
    ws = wb.create_sheet("손님")
    ws["A1"] = "손님 (결제력 / 만족도)"; ws["A1"].font = TITLE_FONT
    hdr = ["손님", "지갑(₩)", "시작만족", "인내심(초)", "식사+/초", "대기-/초", "식사속도", "이동속도", "weight", "시작해금"]
    put_header(ws, hdr)
    r = 4; wallet_rows = []
    fields = ["wallet", "baseSatisfaction", "patience", "eatGainRate", "waitPenaltyRate",
              "eatSpeed", "moveSpeed", "spawnWeight", "unlockedFromStart"]
    for f in sorted(glob.glob(f"{DATAS}/Customer/*.asset")):
        d = load(f, fields)
        name = os.path.splitext(os.path.basename(f))[0]
        ws.cell(r, 1, name).font = BASE_FONT
        if d["wallet"] is not None:
            ws.cell(r, 2, d["wallet"]).number_format = WON; wallet_rows.append(r)
        else:
            ws.cell(r, 2, "(DT/배달)")
        ws.cell(r, 3, d["baseSatisfaction"]); ws.cell(r, 4, d["patience"])
        ws.cell(r, 5, d["eatGainRate"]);      ws.cell(r, 6, d["waitPenaltyRate"])
        ws.cell(r, 7, d["eatSpeed"]);         ws.cell(r, 8, d["moveSpeed"])
        ws.cell(r, 9, d["spawnWeight"])
        ws.cell(r, 10, "" if d["unlockedFromStart"] is None
                       else ("O" if d["unlockedFromStart"] == 1 else "X"))
        for c in range(1, 11):
            ws.cell(r, c).border = BORDER
            if 2 <= c <= 10: ws.cell(r, c).font = INPUT_FONT; ws.cell(r, c).fill = INPUT_FILL
        r += 1
    avg_wallet_cell = None
    ws.cell(r, 1, "평균 지갑(매장)").font = Font(name=ARIAL, bold=True)
    if wallet_rows:
        rng = ",".join(f"B{x}" for x in wallet_rows)
        ws.cell(r, 2, f"=AVERAGE({rng})").number_format = WON
        ws.cell(r, 2).font = Font(name=ARIAL, bold=True)
        avg_wallet_cell = f"손님!B{r}"
    autow(ws, [16, 11, 9, 10, 9, 9, 9, 9, 8, 9])

    # ───────────── 직원 ─────────────
    ws = wb.create_sheet("직원")
    ws["A1"] = "직원 (인건비 / 능력)"; ws["A1"].font = TITLE_FONT
    hdr = ["직원", "고용비(₩)", "월급(₩)", "이동속도", "속도배수", "친절", "배달시간"]
    put_header(ws, hdr)
    order = ["Cook_Junior", "Cook_Senior", "Cook_Manager",
             "Server_Junior", "Server_Senior", "Server_Manager",
             "Rider_Junior", "Rider_Senior", "Rider_Manager"]
    r = 4
    for name in order:
        f = f"{DATAS}/StaffData/{name}.asset"
        if not os.path.exists(f): continue
        d = load(f, ["hireCost", "salary", "moveSpeed", "speedMultiplier", "kindness", "deliveryTime"])
        ws.cell(r, 1, name).font = BASE_FONT
        ws.cell(r, 2, d["hireCost"]).number_format = WON
        ws.cell(r, 3, d["salary"]).number_format = WON
        ws.cell(r, 4, d["moveSpeed"]); ws.cell(r, 5, d["speedMultiplier"])
        ws.cell(r, 6, d["kindness"]);  ws.cell(r, 7, d["deliveryTime"])
        for c in range(1, 8):
            ws.cell(r, c).border = BORDER
            if c >= 2: ws.cell(r, c).font = INPUT_FONT; ws.cell(r, c).fill = INPUT_FILL
        r += 1
    autow(ws, [16, 12, 12, 10, 10, 8, 10])

    # ───────────── 가구 / 조리도구 ─────────────
    ws = wb.create_sheet("가구")
    ws["A1"] = "가구 / 조리도구"; ws["A1"].font = TITLE_FONT
    ws["A2"] = "조리타입은 읽기전용(매칭 enum). 사용시간=조리시간."
    ws["A2"].font = Font(name=ARIAL, italic=True, size=9, color="808080")
    hdr = ["항목", "조리타입", "설치비(₩)", "해금비(만족도)", "사용시간(초)"]
    put_header(ws, hdr)
    furn = (sorted(glob.glob(f"{DATAS}/FurnitureData/*.asset")) +
            sorted(glob.glob(f"{DATAS}/FurnitureData/CookingToolData/*.asset")) +
            sorted(glob.glob(f"{DATAS}/FurnitureData/Toilet/*.asset")))
    r = 4
    for f in furn:
        name = os.path.splitext(os.path.basename(f))[0]
        if name == "LayoutData": continue
        d = load(f, ["purchaseCost", "satisfactionUnlock", "usingDuration", "toolType"])
        ws.cell(r, 1, name).font = BASE_FONT
        ws.cell(r, 2, d["toolType"] if d["toolType"] is not None else "-").fill = CALC_FILL
        ws.cell(r, 3, d["purchaseCost"]).number_format = WON; ws.cell(r, 3).font = INPUT_FONT; ws.cell(r, 3).fill = INPUT_FILL
        ws.cell(r, 4, d["satisfactionUnlock"]).number_format = WON; ws.cell(r, 4).font = INPUT_FONT; ws.cell(r, 4).fill = INPUT_FILL
        ws.cell(r, 5, d["usingDuration"] if d["usingDuration"] is not None else "-")
        if d["usingDuration"] is not None:
            ws.cell(r, 5).font = INPUT_FONT; ws.cell(r, 5).fill = INPUT_FILL
        for c in range(1, 6): ws.cell(r, c).border = BORDER
        r += 1
    autow(ws, [18, 9, 12, 14, 12])

    # ───────────── 맵확장 ─────────────
    ws = wb.create_sheet("맵확장")
    ws["A1"] = "맵 확장 게이트"; ws["A1"].font = TITLE_FONT
    put_header(ws, ["확장", "비용(₩)", "설명"])
    exp = [("Stage1_DTUnlock", "DT(드라이브스루) 해금"),
           ("Stage2_Floor2Hall", "2층 홀 확장"),
           ("Stage3_Floor2Toilet", "2층 화장실 확장")]
    r = 4; gate_cell = {}
    for fn, label in exp:
        p = f"{DATAS}/ExpansionData/{fn}.asset"
        if not os.path.exists(p): continue
        d = load(p, ["unlockCost"])
        ws.cell(r, 1, fn).font = BASE_FONT
        ws.cell(r, 2, d["unlockCost"]).number_format = WON; ws.cell(r, 2).font = INPUT_FONT; ws.cell(r, 2).fill = INPUT_FILL
        ws.cell(r, 3, label).font = Font(name=ARIAL, size=9, color="808080")
        for c in (1, 2, 3): ws.cell(r, c).border = BORDER
        gate_cell[fn] = f"맵확장!B{r}"
        r += 1
    autow(ws, [20, 12, 22])

    # ───────────── 마케팅 ─────────────
    ws = wb.create_sheet("마케팅")
    ws["A1"] = "마케팅 캠페인"; ws["A1"].font = TITLE_FONT
    put_header(ws, ["마케팅", "boost", "만족도비", "기간(달)", "설명"])
    r = 4
    for fn, label in [("FlyerAd", "전단지"), ("SNSAd", "SNS"), ("TVAd", "TV")]:
        p = f"{DATAS}/MarketingData/{fn}.asset"
        if not os.path.exists(p): continue
        d = load(p, ["spawnBoost", "satisfactionCost", "durationMonths"])
        ws.cell(r, 1, fn).font = BASE_FONT
        ws.cell(r, 2, d["spawnBoost"]); ws.cell(r, 3, d["satisfactionCost"]).number_format = WON
        ws.cell(r, 4, d["durationMonths"])
        ws.cell(r, 5, label).font = Font(name=ARIAL, size=9, color="808080")
        for c in range(1, 6):
            ws.cell(r, c).border = BORDER
            if 2 <= c <= 4: ws.cell(r, c).font = INPUT_FONT; ws.cell(r, c).fill = INPUT_FILL
        r += 1
    ws.cell(r + 1, 1, "실효 배수 = 1 + ln(1 + Σboost)").font = Font(name=ARIAL, italic=True, size=9, color="808080")
    autow(ws, [12, 8, 10, 9, 16])

    # ───────────── 모델 (하루 순이익 시뮬레이터) ─────────────
    ws = wb.create_sheet("모델", 0)
    ws["A1"] = "💰 하루 순이익 모델"; ws["A1"].font = TITLE_FONT
    ws["A2"] = "노랑칸(파란글씨)만 바꾸면 전부 자동 계산됩니다"; ws["A2"].font = Font(name=ARIAL, italic=True, size=9, color="808080")

    def lbl(r, txt, bold=False):
        c = ws.cell(r, 1, txt); c.font = Font(name=ARIAL, bold=bold)
    def inp(r, val, fmt=None, note=None):
        c = ws.cell(r, 2, val); c.font = INPUT_FONT; c.fill = INPUT_FILL; c.border = BORDER
        if fmt: c.number_format = fmt
        if note: ws.cell(r, 3, note).font = Font(name=ARIAL, size=9, color="808080")
    def calc(r, formula, fmt=None, link=False, note=None, bold=False):
        c = ws.cell(r, 2, formula); c.font = LINK_FONT if link else CALC_FONT
        if bold: c.font = Font(name=ARIAL, bold=True, color=("008000" if link else "000000"))
        c.border = BORDER
        if fmt: c.number_format = fmt
        if note: ws.cell(r, 3, note).font = Font(name=ARIAL, size=9, color="808080")

    ws.cell(4, 1, "입력 (가정)").font = HDR_FONT; ws.cell(4, 1).fill = HDR_FILL
    for c in (2, 3): ws.cell(4, c).fill = HDR_FILL
    lbl(5, "영업시간 (초/일)");          inp(5, 240, WON, "(24-8)시 × 15초/시간")
    lbl(6, "평균 스폰간격 (초)");         inp(6, 20, None, "(min10+max30)/2")
    lbl(7, "마케팅 배수");               inp(7, 1.0, '0.00', "마케팅 0개=1.0, 3개=약2.03")
    lbl(8, "좌석 수");                  inp(8, 3, None, "동시 수용 손님 수")
    lbl(9, "손님 1명 평균 점유시간 (초)"); inp(9, 30, None, "주문+조리+식사+퇴장")
    lbl(10, "하루 인건비 합계 (₩)");      inp(10, 1900, WON, "고용 직원 월급 합 (1달=1일)")

    ws.cell(12, 1, "계산").font = HDR_FONT; ws.cell(12, 1).fill = HDR_FILL
    for c in (2, 3): ws.cell(12, c).fill = HDR_FILL
    lbl(13, "유효 스폰간격 (초)");   calc(13, "=B6/B7", '0.0', note="스폰간격 ÷ 마케팅배수")
    lbl(14, "이론상 손님/일");       calc(14, "=B5/B13", '0.0', note="좌석 무제한 가정")
    lbl(15, "처리한계 손님/일");     calc(15, "=B8*B5/B9", '0.0', note="좌석×영업시간 ÷ 점유시간")
    lbl(16, "실제 손님/일", True);   calc(16, "=MIN(B14,B15)", '0.0', bold=True, note="둘 중 작은 값(병목)")
    lbl(17, "평균 지갑 (매장)");     calc(17, f"={avg_wallet_cell}", WON, link=True, note="← 손님 시트")
    lbl(18, "평균 마진율");          calc(18, f"={menu_avg_marginpct}", PCT, link=True, note="← 메뉴 시트")
    lbl(19, "손님당 순이익 (₩)");    calc(19, "=B17*B18", WON, note="지갑 전액 결제 가정")
    lbl(20, "하루 총이익 (₩)", True); calc(20, "=B16*B19", WON, bold=True)
    lbl(21, "하루 인건비 (₩)");      calc(21, "=B10", WON, note="입력값")
    lbl(22, "하루 순이익 (₩)", True)
    c = ws.cell(22, 2, "=B20-B21"); c.font = Font(name=ARIAL, bold=True, size=12, color="C00000")
    c.number_format = WON; c.border = BORDER

    ws.cell(24, 1, "성장 게이트까지 걸리는 일수").font = HDR_FONT; ws.cell(24, 1).fill = HDR_FILL
    for c in (2, 3): ws.cell(24, c).fill = HDR_FILL
    rows = [("Stage1_DTUnlock", "DT 해금"), ("Stage2_Floor2Hall", "2층 해금"), ("Stage3_Floor2Toilet", "화장실 해금")]
    rr = 25
    for fn, label in rows:
        ref = gate_cell.get(fn)
        if ref:
            lbl(rr, label); calc(rr, f'=IF($B$22<=0,"적자",{ref}/$B$22)', '0.0', note="비용 ÷ 하루순이익")
            rr += 1
    autow(ws, [26, 16, 34])

    try:
        wb.save(XLSX)
    except PermissionError:
        print("⚠ 저장 실패: Balance_Model.xlsx 가 열려 있습니다. 엑셀을 닫고 다시 실행하세요.")
        return 1
    print(f"✅ export 완료 → {os.path.basename(XLSX)}  (메뉴에 '조리도구' 매칭 컬럼/드롭다운 포함)")
    return 0

# ══════════════════════════════════════════════════════════════════════════
#  IMPORT  (엑셀 → 에셋)
# ══════════════════════════════════════════════════════════════════════════
def do_import(dry):
    from openpyxl import load_workbook
    if not os.path.exists(XLSX):
        print(f"⚠ {XLSX} 가 없습니다. 먼저 'export' 하세요.")
        return 1

    name2guid, _ = tool_maps()
    index = asset_index()
    wb = load_workbook(XLSX, data_only=False)

    # 에셋별로 바꿀 값 모으기 {path: [(field, kind, raw)]}
    pending = {}
    warnings = []
    for sheet, (id_hdr, fmap) in IMPORT_SCHEMA.items():
        if sheet not in wb.sheetnames:
            warnings.append(f"시트 '{sheet}' 없음 (건너뜀)"); continue
        ws = wb[sheet]
        # 헤더 행 찾기
        hdr_row, cols = None, {}
        for rr in range(1, 12):
            row = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(rr, c).value
                if v is not None: row[str(v).strip()] = c
            # 헤더 행 = id 헤더 + 실제 필드 헤더가 함께 있는 행 (제목행 오인 방지)
            if id_hdr in row and (set(row) & set(fmap)):
                hdr_row, cols = rr, row
                break
        if hdr_row is None:
            warnings.append(f"시트 '{sheet}' 에서 헤더 '{id_hdr}' 못 찾음"); continue
        id_col = cols[id_hdr]
        for rr in range(hdr_row + 1, ws.max_row + 1):
            name = ws.cell(rr, id_col).value
            if name is None or str(name).strip() == "": continue
            name = str(name).strip()
            if name not in index: continue   # 평균/소계 등 비-에셋 행 자동 스킵
            path = index[name]
            for hdr, (fld, kind) in fmap.items():
                if hdr not in cols: continue
                raw = ws.cell(rr, cols[hdr]).value
                if raw is None: continue
                if isinstance(raw, str) and raw.strip() in ("", "-", "(DT/배달)"): continue
                pending.setdefault(path, []).append((fld, kind, raw))

    # 적용
    changed_files = 0; changed_fields = 0
    for path, items in pending.items():
        t = read(path)
        diffs = []
        for fld, kind, raw in items:
            old = field(t, fld)
            if kind == "tool":
                tn = str(raw).strip()
                if tn.lower() in ("none", "null", ""):
                    newval = "{fileID: 0}"
                else:
                    g = name2guid.get(tn)
                    if g is None:
                        warnings.append(f"{os.path.basename(path)}: 알 수 없는 조리도구 '{tn}'"); continue
                    newval = "{fileID: 11400000, guid: %s, type: 2}" % g
                oldguid = parse_guid(old)
                if oldguid == name2guid.get(str(raw).strip()): continue  # 동일
            else:
                newval = fmt_num(kind, raw)
                if old is not None and old.strip() == newval: continue   # 동일 → 스킵
            t2, n = set_line(t, fld, newval)
            if n == 0:
                warnings.append(f"{os.path.basename(path)}: 필드 '{fld}' 라인 없음"); continue
            t = t2
            diffs.append(f"    {fld}: {old} → {newval}")
        if diffs:
            changed_files += 1; changed_fields += len(diffs)
            print(f"  {os.path.basename(path)}")
            print("\n".join(diffs))
            if not dry:
                open(path, "w", encoding="utf-8", newline="\n").write(t)

    print("─" * 60)
    if warnings:
        print("⚠ 경고:")
        for w in warnings: print("   -", w)
    mode = "DRY-RUN (쓰지 않음)" if dry else "적용됨"
    print(f"{mode}: 파일 {changed_files}개 / 필드 {changed_fields}개 변경")
    if not dry and changed_files:
        print("→ git diff 로 확인 후 커밋하세요. Unity 에디터는 에셋 reimport 필요.")
    return 0

# ══════════════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser(description="밸런스 양방향 관리 (에셋 ↔ 엑셀)")
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("export", help="에셋 → Balance_Model.xlsx")
    pi = sub.add_parser("import", help="Balance_Model.xlsx → 에셋")
    pi.add_argument("--dry-run", action="store_true", help="미리보기 (파일 안 건드림)")
    a = ap.parse_args()
    if a.cmd == "export":
        return export()
    return do_import(a.dry_run)

if __name__ == "__main__":
    sys.exit(main())
